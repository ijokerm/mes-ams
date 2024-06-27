#!/git/ijokerm
# -*- coding: UTF-8 -*-
"""
@Project ：mes-ams 
@File    ：excel_handler.py
@Author  ：SpringBear
@Date    ：2024/6/12 10:05 
"""
import openpyxl
import ddt
class ExcelHandle:
    def __init__(self,file):
        self.file = file

    def open_excel(self,sheet_name):
        """ 打开Excel，获取sheet"""
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[sheet_name]
        return sheet
    def get_header(self,sheet_name):
        """ 获取表头 """
        wb = self.open_excel(sheet_name)
        header = []
        # 遍历第一行
        for i in wb[1]:
            # 将遍历出来的表头字段加入列表
            header.append(i.value)
        return header

    def read_excel(self, sheet_name):
        """读取所有数据"""
        sheet = self.open_excel(sheet_name)
        rows = list(sheet.rows)
        data = []
        # 遍历从第二行开始的每一行数据
        for row in rows[1:]:
            row_data = []
            # 遍历每一行的每个单元格
            for cell in row:
                row_data.append(cell.value)
                # 通过zip函数将两个列表合并成字典
                data_dict = dict(zip(self.get_header(sheet_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write_excel(file, sheet_name, row, cloumn, data):
        """Excel写入数据"""
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, cloumn).value = data
        wb.save(file)
        wb.close()

if __name__ == "__main__":
    # 以下为测试代码
    excel = ExcelHandle('../data/testcase.xlsx')
    data = excel.read_excel('p1')

    print(data)